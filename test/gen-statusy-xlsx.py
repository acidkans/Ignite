# -*- coding: utf-8 -*-
"""
Generator arkusza z propozycją taksonomii statusów Ignite (planowanie / realizacja).
Analiza stanu obecnego + propozycja — NIE zmienia kodu aplikacji.
Uruchomienie: python test/gen-statusy-xlsx.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'statusy-ignite-propozycja.xlsx')

FONT = 'Arial'
NAVY = '1F3864'
GREY = 'F2F2F2'
BAND = 'FAFAFA'

TITLE_F = Font(name=FONT, size=14, bold=True, color=NAVY)
NOTE_F = Font(name=FONT, size=9, italic=True, color='595959')
HEAD_F = Font(name=FONT, size=10, bold=True, color='FFFFFF')
CELL_F = Font(name=FONT, size=10)
CELL_B = Font(name=FONT, size=10, bold=True)
HEAD_FILL = PatternFill('solid', fgColor=NAVY)
SEC_FILL = PatternFill('solid', fgColor='D9E2F3')
BAND_FILL = PatternFill('solid', fgColor=BAND)
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def sheet(wb, name, title, note, headers, rows, widths, wrap_cols=None, section_rows=None):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws['A1'] = title
    ws['A1'].font = TITLE_F
    ws['A2'] = note
    ws['A2'].font = NOTE_F
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 26
    ws['A2'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    hr = 4
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=hr, column=i, value=h)
        c.font = HEAD_F
        c.fill = HEAD_FILL
        c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        c.border = BORDER
    ws.row_dimensions[hr].height = 30

    for r, row in enumerate(rows, start=hr + 1):
        is_section = section_rows and (r - hr - 1) in section_rows
        for i, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=val)
            c.font = CELL_B if is_section else CELL_F
            c.border = BORDER
            c.alignment = Alignment(
                wrap_text=(wrap_cols is None or i in wrap_cols),
                vertical='top',
            )
            if is_section:
                c.fill = SEC_FILL
            elif (r - hr) % 2 == 0:
                c.fill = BAND_FILL

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = ws.cell(row=hr + 1, column=1)
    ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(headers))}{hr + len(rows)}"
    return ws


wb = Workbook()
wb.remove(wb.active)

# ─────────────────────────────────────────────────────────────── 0. Legenda
sheet(
    wb, '0. Legenda',
    'Statusy w Ignite — analiza stanu obecnego i propozycja rozdziału osi',
    'Dokument analityczny (stan repo: main, 2026-09-02). Nic w kodzie nie zostało zmienione. '
    'Propozycja rozdziela JEDNO dzisiejsze pole WbsNode.status na trzy niezależne osie: '
    'planowanie (decyzja ofertowa), zakup (droga towaru/zlecenia) i wykonanie (robota w terenie).',
    ['Arkusz', 'Co zawiera', 'Do czego służy'],
    [
        ['1. AS-IS pola', 'Inwentaryzacja wszystkich pól statusowych w bazie i we froncie', 'Punkt wyjścia — co dziś istnieje i gdzie'],
        ['2. AS-IS problemy', 'Lista wad obecnego modelu z dowodem w kodzie', 'Uzasadnienie zmiany'],
        ['3. PLANOWANIE liście', 'Proponowany słownik planStatus (Nowe / Zaproponowane / Zaakceptowane / Odrzucone)', 'Etap planu — wszystkie typy liści tak samo'],
        ['4. PLANOWANIE gałęzie', 'Reguła wyliczania statusu gałęzi z sumy statusów dzieci', 'Gałąź nigdy nie edytowalna ręcznie'],
        ['5. REALIZACJA zakup', 'Proponowany słownik purchaseStatus — materiał, sprzęt, usługa obca, nocleg, paliwo', 'Droga TOWARU / ZLECENIA'],
        ['6. REALIZACJA wykonanie', 'Proponowany słownik execStatus — praca, usługa, montaż materiału i sprzętu', 'Droga ROBOTY'],
        ['7. Macierz typ x oś', 'Który typ liścia ma które osie statusów', 'Rozdział typów liści wg pytania'],
        ['8. Przejścia', 'Dozwolone przejścia między statusami każdej osi', 'Walidacja w backendzie'],
        ['9. Migracja kodów', 'Mapowanie każdego dzisiejszego kodu na nowe osie', 'Skrypt migracyjny 1:1'],
        ['10. Model danych', 'Proponowane zmiany w schema.prisma i w warstwie API', 'Zakres wdrożenia'],
        ['11. Decyzje otwarte', 'Pytania wymagające rozstrzygnięcia przed wdrożeniem', 'Do uzupełnienia przez Ciebie'],
        ['', '', ''],
        ['OZNACZENIA', '', ''],
        ['TAK', 'oś dotyczy tego typu liścia', ''],
        ['—', 'oś nie dotyczy (pole zostaje NULL, kolumna w UI pusta)', ''],
        ['wyliczany', 'wartość nie jest zapisywana w bazie — liczona w runtime z dzieci lub z faktów (LeafActual, protokoły odbioru)', ''],
        ['ręczny', 'użytkownik ustawia z dropdowna', ''],
    ],
    [26, 62, 52],
    section_rows={12},
)

# ─────────────────────────────────────────────────────── 1. AS-IS — pola
sheet(
    wb, '1. AS-IS pola',
    'Stan obecny — gdzie w ogóle żyją statusy',
    'Wszystko poniżej to pola typu String bez enuma w Prisma i bez whitelisty w backendzie — '
    'słownik istnieje wyłącznie we froncie, w pięciu kopiach.',
    ['Pole / stała', 'Plik', 'Kody', 'Etykiety', 'Kto ustawia', 'Uwaga'],
    [
        ['WbsNode.status', 'apps/backend/prisma/schema.prisma (@anchor wbs-node-status)',
         'String @default("")', 'brak — sam string',
         'PATCH /wbs-nodes/:id', 'JEDNO pole na plan i na realizację jednocześnie — źródło całego problemu'],
        ['MaterialRequirement.status', 'apps/backend/prisma/schema.prisma (@anchor mat-req-status)',
         'PENDING | CONFIRMED | REJECTED | PROPOSAL', 'j.w.',
         'PATCH /material-requirements/:id', 'Drugie źródło prawdy dla tej samej pozycji'],
        ['STRUCTURE_STATUS_META', 'apps/frontend/src/components/shared/wbs/wbsConstants.js',
         "'' , NEW, PENDING, PROPOSAL, CONFIRMED, REJECTED, ORDERED, EXTRA_ORDER, IN_STOCK, ISSUED, DONE, INSTALLED, MIXED",
         'Brak, Nowy, Oczekuje, Propozycja, Potwierdzone, Odrzucone, Zamówione, Dodatkowe zamówienie, Na magazynie, Wydane, Wykonane, Zainstalowane, Mieszany',
         'dropdown w WBS / Realizacji', 'Słownik "materiałowy" — miesza etap oferty z etapem magazynu'],
        ['WORK_STATUS_META', 'apps/frontend/src/components/shared/wbs/wbsConstants.js',
         'NEW, STARTED, ON_HOLD, COMPLETED, UNFINISHED, CANCELLED',
         'Nowe, Rozpoczęte, Wstrzymane, Zakończone, Nieskończone, Odwołane',
         'dropdown dla work/service/lodging/fuel', 'Słownik "robociznowy" — jedzie tą samą kolumną WbsNode.status'],
        ['STRUCT_STATUS_META / WORK_STRUCT_STATUS_META', 'apps/frontend/src/components/shared/wbs/WBSHybridTable.jsx',
         'kopia obu powyższych w kształcie plakietki', 'j.w.', '—', 'Kopia 3 i 4 tego samego słownika'],
        ['STATUS_META', 'apps/frontend/src/components/shared/wbs/realizationShared.js',
         'NEW, PENDING, PROPOSAL, CONFIRMED, REJECTED, ORDERED, EXTRA_ORDER, IN_STOCK, ISSUED',
         'j.w. (bez DONE / INSTALLED)', '—', 'Kopia 5 — i już niepełna wobec oryginału'],
        ['usesWorkStatuses(type)', 'wbsConstants.js (@anchor uses-work-statuses)',
         "['work','service','lodging','fuel']", '—', 'runtime',
         'Jedyne dzisiejsze rozdzielenie typów liści — po słowniku, nie po etapie'],
        ['resolveStatusCode(type,status)', 'wbsConstants.js (@anchor resolve-status-code)',
         '—', '—', 'runtime',
         'Maskuje brudne dane: kod materiałowy na robociźnie pokazuje jako "Nowe" bez migracji bazy'],
        ['MIXED (gałąź / pozycja materiałowa)', 'UnifiedWbsPanel.jsx ~5731 (getInheritedMaterialStatus)',
         'MIXED', 'Mieszany', 'wyliczany',
         'Jedyna dzisiejsza agregacja — tylko material/equipment i tylko po kluczu nazwa+przedmiot'],
        ['LeafActual (wpisy realizacji)', 'schema.prisma (@anchor leaf-actual-id)',
         'brak statusu — qty, unitCost, docNumber, supplier, data',
         '—', 'użytkownik dopisuje wpis',
         'FAKTY realizacji już są w bazie, ale żaden status się z nich nie wylicza'],
        ['realizationOf().state', 'realizationShared.js (@anchor realization-of)',
         'none | part | full | over | closed', 'kolor paska pokrycia', 'wyliczany',
         'De facto istniejący status realizacji — ale tylko kolor, nie pole i nie do filtrowania'],
        ['WbsNode.realizationClosed', 'schema.prisma (@anchor wbs-node-realization-closed)',
         'Boolean', 'pozycja rozliczona mimo niedowykonania', 'ręcznie PM',
         'Osobna flaga, która powinna być stanem osi realizacji'],
        ['AcceptanceProtocolItem.pelny', 'schema.prisma (@anchor acceptance-protocol-item-pelny)',
         'Boolean', 'pozycja domknięta protokołem odbioru', 'eksport protokołu',
         'Odbiór formalny nigdzie nie wchodzi do statusu pozycji'],
        ['OrderRequirements.offerStatus', 'schema.prisma (linia ~331)',
         'String?', '—', 'ręcznie', 'Status CAŁEJ oferty — inny poziom niż pozycja, zostaje bez zmian'],
        ['Subtask.status', 'schema.prisma (linia ~349)', 'String @default("NEW")', '—', 'ręcznie',
         'Zadania, nie pozycje WBS — poza zakresem tej analizy'],
    ],
    [30, 44, 40, 40, 24, 52],
)

# ──────────────────────────────────────────────────── 2. AS-IS — problemy
sheet(
    wb, '2. AS-IS problemy',
    'Co dokładnie nie działa w obecnym modelu',
    'Każdy problem ma wskazanie miejsca w kodzie — do sprawdzenia przed decyzją o zmianie.',
    ['#', 'Problem', 'Objaw dla użytkownika', 'Dowód w kodzie', 'Rozwiązuje to'],
    [
        [1, 'Jedno pole na dwa etapy życia pozycji',
         'Pozycja "Zamówione" traci informację, że była zaakceptowana; nie da się filtrować "co klient przyjął, a jeszcze nie kupione"',
         'WbsNode.status — jeden String, dwa słowniki wg typu (wbsConstants.js)',
         'Rozdział na planStatus / purchaseStatus / execStatus'],
        [2, 'Statusy planu wymieszane ze statusami magazynu',
         'W dropdownie na etapie wyceny widać "Na magazynie" i "Wydane" — stany, których na planie być nie może',
         'STRUCTURE_STATUS_META — 12 kodów w jednej liście',
         'Arkusze 3, 5 i 6 — trzy rozłączne słowniki'],
        [3, 'Gałąź nie agreguje statusów dzieci',
         'Gałąź "Kamery" pokazuje własny (przypadkowy) status albo nic, mimo że 12 z 20 pozycji pod nią jest zaakceptowanych',
         'Agregacja istnieje TYLKO dla material/equipment i tylko po lookupKey (UnifiedWbsPanel ~5702)',
         'Arkusz 4 — reguła agregacji dla każdej osi i każdego typu'],
        [4, 'Dwa źródła prawdy o statusie tej samej pozycji',
         'Zapis idzie dwoma osobnymi PATCH-ami; gdy drugi padnie, WBS i panel Materiały pokazują różne statusy',
         'RealizationTab.jsx saveStatus (@anchor realization-save-status) — PATCH /wbs-nodes + PATCH /material-requirements',
         'Jedno pole na węźle WBS; karta produktowa status czyta, nie trzyma'],
        [5, 'Brak walidacji kodów w backendzie',
         'Do bazy wchodzi dowolny string; w danych żyją kody materiałowe na robociźnie',
         'schema.prisma: status String @default("") — brak enuma; brak whitelisty w wbs-nodes.service.ts',
         'Enumy Prisma + walidacja DTO'],
        [6, 'Status realizacji trzeba klikać ręcznie, choć fakty są w bazie',
         'Pozycja z pełną dostawą (LeafActual) dalej wisi jako "Zamówione", dopóki ktoś nie przestawi dropdowna',
         'realizationOf() liczy pokrycie, ale wynik idzie tylko na kolor paska',
         'Statusy wyliczane z LeafActual i protokołów (arkusze 5, 6)'],
        [7, 'Brak symetrii kodów między typami',
         'Praca ma "Odwołane", materiał nie ma odpowiednika; materiał ma "Wydane", praca nie ma "Odebrane"',
         'WORK_STATUS_META vs STRUCTURE_STATUS_META',
         'Wspólny szkielet: TO_* / IN_PROGRESS / DONE / CANCELLED w obu osiach realizacji'],
        [8, 'Pięć kopii tego samego słownika we froncie',
         'Zmiana etykiety w jednym miejscu nie propaguje się do eksportu / drugiego widoku',
         'wbsConstants.js x2, WBSHybridTable.jsx x2, realizationShared.js x1',
         'Jeden moduł słownika + typy wyprowadzone z enuma backendu'],
        [9, 'Odrzucone pozycje liczą się jak każde inne',
         'Nie da się prosto policzyć wartości zakresu bez pozycji odrzuconych przez klienta',
         'Brak reguły wykluczania REJECTED z agregatów',
         'Arkusz 4 — REJECTED wypada z podstawy agregacji'],
        [10, 'Brak śladu decyzji (kto / kiedy / dlaczego)',
         'Zmiana z "Zaproponowane" na "Odrzucone" nie zostawia autora ani daty',
         'WbsNode nie ma pól statusChangedBy/At; AuditLog nie obejmuje statusów pozycji',
         'Pola statusChangedAt / statusChangedBy albo wpis do AuditLog'],
    ],
    [5, 40, 52, 56, 44],
)

# ────────────────────────────────────────────── 3. PLANOWANIE — liście
sheet(
    wb, '3. PLANOWANIE liście',
    'ETAP PLANU — planStatus (pozycja OFERTOWA, nie rzecz realna)',
    'Dotyczy WSZYSTKICH typów liści jednakowo — na etapie planu nie ma znaczenia, czy pozycja to switch, '
    'czy dzień pracy ekipy: to wiersz oferty, który klient przyjmie albo nie. Edytowalny WYŁĄCZNIE na liściu.',
    ['Kod', 'Etykieta', 'Znaczenie', 'Kto / co ustawia', 'Wchodzi do zakresu realizacji?',
     'Wchodzi do wartości oferty?', 'Kolor (propozycja)', 'Status'],
    [
        ['NEW', 'Nowe',
         'Pozycja powstała w WBS. Nie poszła jeszcze do klienta — zakres i wycena w robocie.',
         'automat przy utworzeniu węzła', 'NIE', 'TAK (wersja robocza)', 'szary (slate-300)', 'podstawowy'],
        ['PROPOSED', 'Zaproponowane',
         'Pozycja weszła do oferty wysłanej klientowi. Czekamy na decyzję.',
         'automat przy eksporcie oferty (PDF/Excel) albo ręcznie PM', 'NIE', 'TAK', 'niebieski (blue-400)', 'podstawowy'],
        ['ACCEPTED', 'Zaakceptowane',
         'Klient przyjął pozycję. Dopiero teraz pozycja istnieje dla realizacji — zakupów i robót.',
         'ręcznie PM po decyzji klienta (docelowo: akceptacja wersji oferty hurtem)', 'TAK', 'TAK', 'zielony (emerald-400)', 'podstawowy'],
        ['REJECTED', 'Odrzucone',
         'Klient nie przyjął pozycji. Wiersz zostaje w drzewie jako ślad, ale wypada z zakresu, z budżetu i z agregatów gałęzi.',
         'ręcznie PM po decyzji klienta', 'NIE', 'NIE', 'czerwony (red-400)', 'podstawowy'],
        ['WITHDRAWN', 'Wycofane',
         'Pozycja usunięta z zakresu przez NAS przed wysłaniem oferty (dubel, błąd, zmiana koncepcji). '
         'Różni się od "Odrzucone": klient jej nigdy nie widział.',
         'ręcznie PM', 'NIE', 'NIE', 'ciemnoszary (gray-500)', 'OPCJONALNY — do decyzji'],
        ['OPTIONAL', 'Opcja / warunkowe',
         'Pozycja pokazana klientowi jako opcja dodatkowa. Decyzja odroczona — nie blokuje zamknięcia oferty.',
         'ręcznie PM', 'NIE (do czasu decyzji)', 'osobna suma "opcje"', 'fioletowy (violet-400)', 'OPCJONALNY — do decyzji'],
        ['', '', 'ZASADY OSI PLANOWANIA', '', '', '', '', ''],
        ['R1', '—', 'Status planowania ustawia się TYLKO na liściu. Gałąź i przedmiot projektu mają wartość wyliczaną (arkusz 4).',
         '', '', '', '', 'zasada'],
        ['R2', '—', 'Stan startowy każdej nowej pozycji to NEW — niezależnie od typu (dziś tak samo: DEFAULT_STATUS_NEW).',
         '', '', '', '', 'zasada'],
        ['R3', '—', 'Przejście do osi realizacji otwiera WYŁĄCZNIE ACCEPTED. Pozycja NEW/PROPOSED/REJECTED nie ma statusu zakupu ani wykonania (NULL).',
         '', '', '', '', 'zasada'],
        ['R4', '—', 'Zmiana z ACCEPTED na REJECTED przy istniejących wpisach realizacji (LeafActual) wymaga potwierdzenia — koszty już poniesiono.',
         '', '', '', '', 'zasada'],
        ['R5', '—', 'Status planowania NIE zmienia się sam z faktów realizacji. To decyzja handlowa, nie stan magazynu.',
         '', '', '', '', 'zasada'],
    ],
    [14, 20, 62, 40, 22, 20, 22, 20],
    section_rows={6},
)

# ─────────────────────────────────────────── 4. PLANOWANIE — gałęzie
sheet(
    wb, '4. PLANOWANIE gałęzie',
    'Gałąź bierze status z sumy statusów swoich dzieci — reguła',
    'Ta sama mechanika obowiązuje w każdej z trzech osi (plan / zakup / wykonanie). '
    'Wartość NIE jest zapisywana w bazie — liczona w runtime, dokładnie jak dziś depth w buildDepths().',
    ['Lp.', 'Warunek na zbiorze liści poddrzewa', 'Status gałęzi', 'Etykieta', 'Uzasadnienie'],
    [
        [1, 'Poddrzewo nie ma żadnego liścia kosztowego', '—', 'Brak', 'Nie ma z czego liczyć; gałąź czysto porządkowa'],
        [2, 'Wszystkie liście = REJECTED', 'REJECTED', 'Odrzucone', 'Cała gałąź wypadła z zakresu'],
        [3, 'Po pominięciu REJECTED: wszystkie = ACCEPTED', 'ACCEPTED', 'Zaakceptowane', 'Gałąź w całości w zakresie'],
        [4, 'Po pominięciu REJECTED: wszystkie = NEW', 'NEW', 'Nowe', 'Nic jeszcze nie poszło do klienta'],
        [5, 'Po pominięciu REJECTED: wszystkie = PROPOSED', 'PROPOSED', 'Zaproponowane', 'Cała gałąź czeka na decyzję'],
        [6, 'Po pominięciu REJECTED: mieszanka ZAWIERAJĄCA ACCEPTED', 'PARTIAL', 'Częściowo zaakceptowane',
         'Najbardziej informatywny przypadek — klient przyjął część zakresu'],
        [7, 'Po pominięciu REJECTED: mieszanka NEW + PROPOSED (bez ACCEPTED)', 'NEW', 'Nowe',
         'Zasada najsłabszego ogniwa: gałąź nie jest "zaproponowana", dopóki choć jedna pozycja nie jest gotowa'],
        ['', 'MIARY TOWARZYSZĄCE (obok plakietki, nie zamiast niej)', '', '', ''],
        ['M1', 'Licznik pozycji: ile liści zaakceptowanych / ile w zakresie', 'np. 12/20', '—',
         'Plakietka mówi "częściowo", licznik mówi "jak bardzo"'],
        ['M2', 'Udział wartościowy: suma totalPrice liści ACCEPTED / suma totalPrice liści w zakresie',
         'np. 78%', '—', 'Dwie pozycje odrzucone potrafią być połową wartości gałęzi'],
        ['', 'WARIANT ALTERNATYWNY (prostszy, do rozważenia)', '', '', ''],
        ['A1', 'Każda niejednorodność zbioru', 'MIXED', 'Mieszany',
         'Tak działa dziś agregacja materiałowa. Prostsze do wdrożenia, ale "Mieszany" nie odróżnia '
         '"prawie wszystko przyjęte" od "prawie nic"'],
        ['', 'ZASADY WSPÓLNE', '', '', ''],
        ['R6', 'Gałąź jest zawsze READ-ONLY — dropdown zablokowany, plakietka z ikoną dziedziczenia', '', '', 'Jak dzisiejszy InheritedStatusBadge'],
        ['R7', 'Agregacja liczy się REKURENCYJNIE po liściach całego poddrzewa, nie po bezpośrednich dzieciach', '', '', 'Inaczej gałąź 3 poziomy wyżej gubi informację'],
        ['R8', 'Węzeł z typem kosztowym, który MA dzieci, liczy się jako liść (niesie własny koszt)', '', '', 'Wprost reguła z leafNodesOf() w realizationShared.js'],
        ['R9', 'Zmiana statusu liścia przelicza gałęzie natychmiast w UI, bez zapisu do bazy', '', '', 'Brak dodatkowych kolumn i brak ryzyka rozjazdu'],
    ],
    [7, 62, 20, 26, 62],
    section_rows={7, 10, 12},
)

# ────────────────────────────────────────────── 5. REALIZACJA — zakup
sheet(
    wb, '5. REALIZACJA zakup',
    'ETAP REALIZACJI, oś ZAKUP — purchaseStatus (droga TOWARU lub ZLECENIA)',
    'Dotyczy tego, co kupujemy na zewnątrz: materiał, sprzęt, usługa obca (podwykonawca), nocleg, paliwo. '
    'Praca własna tej osi NIE MA. Otwiera się dopiero, gdy planStatus = ACCEPTED.',
    ['Kod', 'Etykieta', 'Znaczenie', 'Typy liści', 'Ręczny / wyliczany', 'Wyzwalacz (fakt w bazie)', 'Kolor (propozycja)', 'Status'],
    [
        ['TO_ORDER', 'Do zamówienia',
         'Pozycja w zakresie, zakup jeszcze nie ruszył.', 'materiał, sprzęt, usługa, nocleg, paliwo',
         'automat', 'planStatus przeszedł na ACCEPTED', 'szary (slate-300)', 'podstawowy'],
        ['RFQ', 'Zapytanie u dostawcy',
         'Wysłano zapytanie ofertowe, czekamy na cenę / potwierdzenie dostępności. '
         'To jest prawdziwe znaczenie dzisiejszego "Oczekuje".', 'materiał, sprzęt, usługa',
         'ręczny (logistyk)', '—', 'bursztyn (amber-400)', 'podstawowy'],
        ['ORDERED', 'Zamówione',
         'Zamówienie złożone u dostawcy — jest numer i termin.', 'wszystkie zakupowe',
         'ręczny + podpowiedź', 'LeafActual z docNumber typu zamówienie', 'fiolet (violet-400)', 'podstawowy'],
        ['PARTIALLY_DELIVERED', 'Dostawa częściowa',
         'Dotarła część zamówionej ilości.', 'materiał, sprzęt',
         'WYLICZANY', '0 < suma LeafActual.qty < WbsNode.quantity', 'bursztyn (amber-300)', 'podstawowy'],
        ['DELIVERED', 'Dostarczone / na magazynie',
         'Pełna ilość dotarła i leży na magazynie.', 'materiał, sprzęt',
         'WYLICZANY', 'suma LeafActual.qty >= WbsNode.quantity', 'cyan (cyan-400)', 'podstawowy'],
        ['ISSUED', 'Wydane na budowę',
         'Towar wydany z magazynu ekipie. Ostatni stan, w którym pozycja jest jeszcze "towarem".',
         'materiał, sprzęt', 'ręczny (magazyn)', '—', 'zielony (emerald-400)', 'podstawowy'],
        ['INVOICED', 'Zafakturowane',
         'Faktura zakupu wpięta — koszt rzeczywisty pozycji domknięty.', 'wszystkie zakupowe',
         'ręczny + podpowiedź', 'LeafActual.docNumber = nr FV', 'teal (teal-400)', 'podstawowy'],
        ['CANCELLED', 'Zakup anulowany',
         'Zamówienie odwołane — pozycja nie zostanie kupiona (zmiana zakresu, zamiennik, rezygnacja).',
         'wszystkie zakupowe', 'ręczny', '—', 'czerwony (red-400)', 'podstawowy'],
        ['CLOSED_SHORT', 'Zamknięte niedoborem',
         'Pozycja rozliczona mimo niedostarczenia całej ilości — różnica przestaje być brakiem i liczy się jako oszczędność.',
         'wszystkie zakupowe', 'ręczny (PM)', 'dzisiejsze WbsNode.realizationClosed = true', 'teal (teal-300)', 'podstawowy'],
        ['', '', 'ZNACZNIKI — nie są stanem, doklejają się do stanu', '', '', '', '', ''],
        ['~extra', 'Domówienie',
         'Zakup ponad plan: brak w dostawie, uszkodzenie, zmiana zakresu na budowie. Dziś jest to STAN (EXTRA_ORDER), '
         'a powinien być znacznikiem — pozycja jednocześnie bywa "Dostarczona" i "domówiona".',
         'materiał, sprzęt', 'WYLICZANY', 'suma LeafActual.qty > WbsNode.quantity', 'fuksja (fuchsia-400)', 'znacznik'],
        ['~multi', 'Kilka cen zakupu',
         'Pozycja kupowana w kilku transzach po różnych cenach — istotne przy rozliczeniu.', 'materiał, sprzęt',
         'WYLICZANY', 'realizationOf().mixedPrices = true', 'szary', 'znacznik'],
        ['', '', 'ZASADY OSI ZAKUPU', '', '', '', '', ''],
        ['R10', '—', 'Pole NULL dla liścia typu "praca" — kolumna w UI pusta, nie "Nie dotyczy" w bazie.',
         '', '', '', '', 'zasada'],
        ['R11', '—', 'Stany wyliczane (PARTIALLY_DELIVERED / DELIVERED) nadpisują ręczny ORDERED — fakt dostawy jest silniejszy niż deklaracja.',
         '', '', '', '', 'zasada'],
        ['R12', '—', 'ISSUED i INVOICED są ręczne, bo nie wynikają z żadnego dzisiejszego faktu w bazie.',
         '', '', '', '', 'zasada'],
        ['R13', '—', 'Usługa obca: DELIVERED = podwykonawca zszedł z budowy z fakturą; sam odbiór robót jest na osi WYKONANIE.',
         '', '', '', '', 'zasada'],
    ],
    [20, 24, 60, 26, 20, 34, 22, 16],
    section_rows={9, 12},
)

# ────────────────────────────────────────── 6. REALIZACJA — wykonanie
sheet(
    wb, '6. REALIZACJA wykonanie',
    'ETAP REALIZACJI, oś WYKONANIE — execStatus (droga ROBOTY w terenie)',
    'Dotyczy tego, co ktoś fizycznie robi: praca własna, usługa obca oraz MONTAŻ materiału i sprzętu. '
    'Nocleg i paliwo tej osi NIE MAJĄ — nie ma tam czego wykonywać. Otwiera się przy planStatus = ACCEPTED.',
    ['Kod', 'Etykieta (praca / usługa)', 'Etykieta (materiał / sprzęt)', 'Znaczenie', 'Ręczny / wyliczany',
     'Wyzwalacz (fakt w bazie)', 'Kolor (propozycja)', 'Status'],
    [
        ['TO_DO', 'Do wykonania', 'Do montażu',
         'Pozycja w zakresie, roboty nie ruszyły.', 'automat', 'planStatus przeszedł na ACCEPTED', 'szary (slate-300)', 'podstawowy'],
        ['IN_PROGRESS', 'W toku', 'Montaż w toku',
         'Ekipa rozpoczęła. Dzisiejsze "Rozpoczęte".', 'ręczny + podpowiedź',
         'pierwszy wpis LeafActual na pozycji', 'niebieski (blue-400)', 'podstawowy'],
        ['ON_HOLD', 'Wstrzymane', 'Montaż wstrzymany',
         'Przerwane z przyczyny zewnętrznej: brak frontu robót, brak materiału, decyzja klienta.',
         'ręczny', '—', 'bursztyn (amber-400)', 'podstawowy'],
        ['DONE', 'Wykonane', 'Zainstalowane',
         'Roboty zakończone zgodnie z planem. Dwie etykiety, jeden kod — dokładnie jak dziś NEW ma "Nowy"/"Nowe".',
         'ręczny + podpowiedź', 'suma LeafActual.qty >= WbsNode.quantity', 'zielony (emerald-400)', 'podstawowy'],
        ['HANDED_OVER', 'Odebrane', 'Odebrane',
         'Podpisany protokół odbioru obejmuje tę pozycję — stan formalny, nie techniczny.',
         'WYLICZANY', 'AcceptanceProtocolItem na wbsRootId pozycji', 'lime (lime-400)', 'podstawowy'],
        ['UNFINISHED', 'Niedokończone', 'Montaż niedokończony',
         'Przerwane przed metą i nie będzie kontynuowane; rozliczamy tyle, ile zrobiono. Dzisiejsze "Nieskończone".',
         'ręczny (PM)', 'WbsNode.realizationClosed przy pokryciu < 100%', 'pomarańcz (orange-400)', 'podstawowy'],
        ['CANCELLED', 'Odwołane', 'Montaż odwołany',
         'Nigdy nie ruszyło i nie ruszy — pozycja wypada z realizacji bez kosztu.', 'ręczny', '—', 'czerwony (red-400)', 'podstawowy'],
        ['', '', '', 'ZNACZNIKI', '', '', '', ''],
        ['~over', 'Ponad plan', 'Ponad plan',
         'Wykonano więcej, niż zaplanowano (dzisiejszy state = over w realizationOf).', 'WYLICZANY',
         'suma LeafActual.qty > WbsNode.quantity', 'czerwony', 'znacznik'],
        ['~partial_accept', 'Odbiór częściowy', 'Odbiór częściowy',
         'Protokół objął część wartości pozycji (AcceptanceProtocolItem.pelny = false).', 'WYLICZANY',
         'suma wartości protokołów < totalPrice', 'lime jasny', 'znacznik'],
        ['', '', '', 'ZASADY OSI WYKONANIA', '', '', '', ''],
        ['R14', '—', '—', 'Pole NULL dla noclegu i paliwa.', '', '', '', 'zasada'],
        ['R15', '—', '—', 'Materiał i sprzęt mają OBIE osie naraz: kupione (ISSUED) to jeszcze nie zamontowane (DONE). '
         'Dziś jedno pole nie potrafi tego powiedzieć.', '', '', '', 'zasada'],
        ['R16', '—', '—', 'HANDED_OVER jest zawsze PO DONE — odbiór nie może wyprzedzić wykonania.', '', '', '', 'zasada'],
        ['R17', '—', '—', 'Pozycja "Do wykonania" z pełną dostawą to sygnał dla ekipy — najważniejszy filtr dnia w terenie.',
         '', '', '', 'zasada'],
    ],
    [16, 24, 24, 56, 20, 34, 22, 16],
    section_rows={7, 10},
)

# ──────────────────────────────────────────────── 7. Macierz typ x oś
sheet(
    wb, '7. Macierz typ x oś',
    'Który typ liścia ma które osie statusów',
    'Rozdział typów liści, o który pytasz: materiał i sprzęt idą OBIEMA osiami realizacji, '
    'praca tylko wykonaniem, usługa obca obiema, nocleg i paliwo tylko zakupem.',
    ['Typ liścia (WbsNode.type)', 'Etykieta PL', 'Ma kartę produktową', 'planStatus', 'purchaseStatus',
     'execStatus', 'Dlaczego tak'],
    [
        ['material', 'Materiał', 'TAK (MaterialRequirement)', 'TAK', 'TAK', 'TAK',
         'Kupujemy rzecz, a potem ją montujemy — dwa niezależne przebiegi na jednej pozycji'],
        ['equipment', 'Sprzęt', 'TAK (MaterialRequirement)', 'TAK', 'TAK', 'TAK', 'Jak materiał'],
        ['work', 'Praca', 'NIE', 'TAK', '— (NULL)', 'TAK',
         'Robocizna własna: nie ma czego zamawiać ani wydawać z magazynu'],
        ['service', 'Usługa obca', 'NIE', 'TAK', 'TAK', 'TAK',
         'Podwykonawcę się ZAMAWIA (zlecenie, umowa, faktura) i osobno ODBIERA jego robotę'],
        ['lodging', 'Nocleg', 'NIE', 'TAK', 'TAK', '— (NULL)',
         'Koszt nabywany na zewnątrz; nic nie jest "wykonywane" w terenie'],
        ['fuel', 'Paliwo', 'NIE', 'TAK', 'TAK', '— (NULL)', 'Jak nocleg'],
        ['group', 'Grupujący', 'NIE', 'wyliczany', 'wyliczany', 'wyliczany', 'Gałąź — reguła z arkusza 4'],
        ['(pusty)', 'Bez typu', 'NIE', 'TAK', '— (NULL)', '— (NULL)',
         'Nie wiadomo, czym pozycja jest — do czasu nadania typu tylko oś planu'],
        ['', '', '', '', '', '', ''],
        ['UWAGA', '', '', '', '', '', ''],
        ['Dzisiejszy podział', 'usesWorkStatuses() dzieli na [work, service, lodging, fuel] vs reszta',
         '', '', '', '',
         'Podział jest po SŁOWNIKU, nie po etapie — dlatego nocleg dostaje "Rozpoczęte"/"Zakończone", '
         'choć nocleg się kupuje, a nie wykonuje'],
        ['Propozycja', 'Podział po OSI: co kupujemy vs co wykonujemy', '', '', '', '',
         'Nocleg i paliwo trafiają tam, gdzie ich miejsce — na oś zakupu'],
    ],
    [24, 16, 26, 14, 16, 14, 62],
    section_rows={9},
)

# ───────────────────────────────────────────────────── 8. Przejścia
sheet(
    wb, '8. Przejścia',
    'Dozwolone przejścia między statusami — materiał do walidacji w backendzie',
    'Kolumna "Dozwolone z" opisuje, z jakiego stanu wolno wejść do danego. Wszystko poza tą listą backend odrzuca — '
    'dziś nie odrzuca niczego, bo pole to goły String.',
    ['Oś', 'Status docelowy', 'Dozwolone z', 'Kto ma prawo', 'Efekt uboczny'],
    [
        ['PLAN', 'NEW', '(utworzenie węzła), PROPOSED', 'PM, projektant', 'Czyści purchaseStatus i execStatus na NULL'],
        ['PLAN', 'PROPOSED', 'NEW, OPTIONAL', 'PM', 'Znacznik wersji oferty, w której pozycja poszła do klienta'],
        ['PLAN', 'ACCEPTED', 'PROPOSED, OPTIONAL, REJECTED', 'PM', 'OTWIERA osie realizacji: purchaseStatus = TO_ORDER, execStatus = TO_DO (wg typu)'],
        ['PLAN', 'REJECTED', 'NEW, PROPOSED, OPTIONAL, ACCEPTED', 'PM',
         'Wypada z agregatów gałęzi i z wartości oferty. Z ACCEPTED tylko po potwierdzeniu, gdy są wpisy realizacji'],
        ['PLAN', 'WITHDRAWN (opcja)', 'NEW', 'PM, projektant', 'Pozycja znika z eksportów, zostaje w drzewie'],
        ['PLAN', 'OPTIONAL (opcja)', 'NEW, PROPOSED', 'PM', 'Osobna suma "opcje" w podsumowaniu oferty'],
        ['', '', '', '', ''],
        ['ZAKUP', 'TO_ORDER', '(automat przy ACCEPTED), RFQ, CANCELLED', 'automat, logistyk', '—'],
        ['ZAKUP', 'RFQ', 'TO_ORDER', 'logistyk', '—'],
        ['ZAKUP', 'ORDERED', 'TO_ORDER, RFQ', 'logistyk', 'Wymaga numeru zamówienia i terminu'],
        ['ZAKUP', 'PARTIALLY_DELIVERED', 'ORDERED', 'wyliczany', 'Nie do ustawienia ręcznie'],
        ['ZAKUP', 'DELIVERED', 'ORDERED, PARTIALLY_DELIVERED', 'wyliczany', 'Nie do ustawienia ręcznie'],
        ['ZAKUP', 'ISSUED', 'DELIVERED, PARTIALLY_DELIVERED', 'magazyn', 'Odblokowuje montaż — sygnał dla ekipy'],
        ['ZAKUP', 'INVOICED', 'DELIVERED, ISSUED', 'logistyk, księgowość', 'Domyka koszt rzeczywisty pozycji'],
        ['ZAKUP', 'CANCELLED', 'TO_ORDER, RFQ, ORDERED', 'logistyk, PM', 'Zablokowane, gdy istnieją wpisy LeafActual'],
        ['ZAKUP', 'CLOSED_SHORT', 'ORDERED, PARTIALLY_DELIVERED', 'PM', 'Ustawia realizationClosed = true; różnica idzie w oszczędność'],
        ['', '', '', '', ''],
        ['WYKONANIE', 'TO_DO', '(automat przy ACCEPTED), ON_HOLD', 'automat, kierownik robót', '—'],
        ['WYKONANIE', 'IN_PROGRESS', 'TO_DO, ON_HOLD', 'kierownik robót, ekipa', 'Podpowiadane przy pierwszym wpisie LeafActual'],
        ['WYKONANIE', 'ON_HOLD', 'IN_PROGRESS', 'kierownik robót', 'Wymaga powodu wstrzymania (komentarz)'],
        ['WYKONANIE', 'DONE', 'IN_PROGRESS, ON_HOLD', 'kierownik robót', 'Podpowiadane przy pokryciu >= 100%'],
        ['WYKONANIE', 'HANDED_OVER', 'DONE', 'wyliczany (protokół)', 'Nie do ustawienia ręcznie — wynika z AcceptanceProtocolItem'],
        ['WYKONANIE', 'UNFINISHED', 'IN_PROGRESS, ON_HOLD', 'PM', 'Ustawia realizationClosed = true'],
        ['WYKONANIE', 'CANCELLED', 'TO_DO, ON_HOLD', 'PM', 'Zablokowane, gdy istnieją wpisy LeafActual'],
    ],
    [14, 26, 44, 26, 62],
    section_rows={6, 16},
)

# ────────────────────────────────────────────────── 9. Migracja kodów
sheet(
    wb, '9. Migracja kodów',
    'Mapowanie każdego dzisiejszego kodu na nowe osie',
    'Kolumna źródłowa to WbsNode.status (String) oraz MaterialRequirement.status. '
    'Mapowanie jest deterministyczne — jeden przebieg UPDATE, bez utraty informacji.',
    ['Źródło', 'Dzisiejszy kod', 'Dzisiejsza etykieta', 'planStatus', 'purchaseStatus', 'execStatus',
     'Znacznik / flaga', 'Uwaga do migracji'],
    [
        ['WbsNode.status', "'' (pusty)", 'Brak', 'NEW', 'NULL', 'NULL', '—', 'Domyślna wartość dzisiejszego pola'],
        ['WbsNode.status', 'NEW', 'Nowy / Nowe', 'NEW', 'NULL', 'NULL', '—', '1:1'],
        ['WbsNode.status', 'PENDING', 'Oczekuje', 'NEW', 'NULL', 'NULL', '—',
         'UWAGA: "Oczekuje" znaczyło "czeka na ofertę dostawcy". Na etapie planu to wciąż NEW; '
         'kto chce zachować sens, migruje na planStatus=ACCEPTED + purchaseStatus=RFQ, ale tylko dla pozycji z zamówień w realizacji'],
        ['WbsNode.status', 'PROPOSAL', 'Propozycja', 'PROPOSED', 'NULL', 'NULL', '—', '1:1'],
        ['WbsNode.status', 'CONFIRMED', 'Potwierdzone', 'ACCEPTED', 'TO_ORDER', 'TO_DO', '—', 'Otwiera obie osie wg typu liścia'],
        ['WbsNode.status', 'REJECTED', 'Odrzucone', 'REJECTED', 'NULL', 'NULL', '—', '1:1'],
        ['WbsNode.status', 'ORDERED', 'Zamówione', 'ACCEPTED', 'ORDERED', 'TO_DO', '—', 'Zamówione = klient wcześniej przyjął pozycję'],
        ['WbsNode.status', 'EXTRA_ORDER', 'Dodatkowe zamówienie', 'ACCEPTED', 'ORDERED', 'TO_DO', '~extra',
         'Stan zamienia się w znacznik — dziś kod gubi informację, w jakiej fazie zakupu pozycja naprawdę jest'],
        ['WbsNode.status', 'IN_STOCK', 'Na magazynie', 'ACCEPTED', 'DELIVERED', 'TO_DO', '—', 'Po migracji stan wyliczany z LeafActual'],
        ['WbsNode.status', 'ISSUED', 'Wydane', 'ACCEPTED', 'ISSUED', 'TO_DO', '—', '1:1'],
        ['WbsNode.status', 'DONE', 'Wykonane', 'ACCEPTED', 'DELIVERED (mat./sprzęt) lub NULL', 'DONE', '—',
         'Dziś jeden kod na dwie osie — po rozdziale trzeba dopisać oś zakupu wg typu'],
        ['WbsNode.status', 'INSTALLED', 'Zainstalowane', 'ACCEPTED', 'ISSUED', 'DONE', '—',
         'Zainstalowane implikuje, że towar wyszedł z magazynu'],
        ['WbsNode.status', 'MIXED', 'Mieszany', 'wyliczany', 'wyliczany', 'wyliczany', '—',
         'Nie migruje — to wartość runtime, nigdy nie powinna trafić do bazy'],
        ['WbsNode.status', 'STARTED', 'Rozpoczęte', 'ACCEPTED', 'NULL (praca) / bez zmian', 'IN_PROGRESS', '—', 'Słownik robociznowy'],
        ['WbsNode.status', 'ON_HOLD', 'Wstrzymane', 'ACCEPTED', 'bez zmian', 'ON_HOLD', '—', '1:1'],
        ['WbsNode.status', 'COMPLETED', 'Zakończone', 'ACCEPTED', 'bez zmian', 'DONE', '—', 'Scala się z materiałowym DONE — jeden kod na obu ścieżkach'],
        ['WbsNode.status', 'UNFINISHED', 'Nieskończone', 'ACCEPTED', 'bez zmian', 'UNFINISHED', '—', '1:1'],
        ['WbsNode.status', 'CANCELLED', 'Odwołane', 'ACCEPTED', 'CANCELLED', 'CANCELLED', '—',
         'Plan zostaje ACCEPTED: klient pozycję przyjął, to realizacja została odwołana'],
        ['', '', '', '', '', '', '', ''],
        ['MaterialRequirement.status', 'PENDING', 'Oczekuje', 'NEW', 'NULL', 'NULL', '—',
         'Pole przestaje być źródłem prawdy — status czytany z węzła WBS'],
        ['MaterialRequirement.status', 'PROPOSAL', 'Propozycja', 'PROPOSED', 'NULL', 'NULL', '—', 'j.w.'],
        ['MaterialRequirement.status', 'CONFIRMED', 'Potwierdzone', 'ACCEPTED', 'TO_ORDER', 'TO_DO', '—', 'j.w.'],
        ['MaterialRequirement.status', 'REJECTED', 'Odrzucone', 'REJECTED', 'NULL', 'NULL', '—', 'j.w.'],
        ['', '', '', '', '', '', '', ''],
        ['WbsNode.realizationClosed', 'true', 'rozliczone mimo niedowykonania', 'ACCEPTED',
         'CLOSED_SHORT (gdy braki w dostawie)', 'UNFINISHED (gdy braki w robociźnie)', '—',
         'Flaga staje się stanem osi; kolumnę można zostawić jako zapasową'],
    ],
    [26, 18, 22, 16, 26, 22, 14, 62],
    section_rows={19, 24},
)

# ────────────────────────────────────────────────── 10. Model danych
sheet(
    wb, '10. Model danych',
    'Propozycja zmian w schema.prisma i w API — zakres wdrożenia',
    'Nic z tego nie zostało wykonane. Lista opisuje, co trzeba by ruszyć, gdyby propozycja została przyjęta.',
    ['Warstwa', 'Element', 'Zmiana', 'Uzasadnienie', 'Ryzyko / uwaga'],
    [
        ['schema.prisma', 'WbsNode.planStatus', 'NOWE pole: enum WbsPlanStatus @default(NEW)',
         'Etap planu — jedna oś dla wszystkich typów liści', 'Wymaga migracji danych z WbsNode.status (arkusz 9)'],
        ['schema.prisma', 'WbsNode.purchaseStatus', 'NOWE pole: enum WbsPurchaseStatus? (NULL = oś nie dotyczy)',
         'Droga towaru / zlecenia', 'NULL dla pracy własnej — nie mylić z brakiem danych'],
        ['schema.prisma', 'WbsNode.execStatus', 'NOWE pole: enum WbsExecStatus? (NULL = oś nie dotyczy)',
         'Droga roboty', 'NULL dla noclegu i paliwa'],
        ['schema.prisma', 'WbsNode.status', 'ZOSTAJE do czasu migracji, potem usunięcie',
         'Bezpieczny rollback przez jeden cykl wersji', 'Nie czytać go po wdrożeniu w żadnym widoku'],
        ['schema.prisma', 'WbsNode.statusChangedAt / statusChangedBy', 'NOWE pola (opcjonalne)',
         'Ślad decyzji: kto i kiedy przestawił status planu', 'Alternatywa: wpis do AuditLog zamiast dwóch kolumn'],
        ['schema.prisma', 'MaterialRequirement.status', 'DEPRECATED — czytane z węzła WBS',
         'Likwiduje drugie źródło prawdy i podwójny PATCH', 'Sprawdzić eksporty i import z arkusza, które to pole czytają'],
        ['schema.prisma', 'enum WbsPlanStatus / WbsPurchaseStatus / WbsExecStatus', 'NOWE enumy',
         'Baza przestaje przyjmować dowolny string', 'Enum w Postgres — dodanie wartości wymaga migracji'],
        ['backend', 'wbs-nodes.service.ts getUnifiedTree()', 'Dokłada statusy WYLICZANE dla gałęzi i pochodne z LeafActual',
         'Ta sama mechanika co buildDepths() — runtime, nie kolumna', 'Koszt: jedno przejście po drzewie + agregacja wpisów'],
        ['backend', 'PATCH /wbs-nodes/:id', 'Walidacja przejść z arkusza 8 + blokada zapisu statusu na gałęzi',
         'Dziś przechodzi każdy string i każde przejście', 'Trzeba przewidzieć ścieżkę dla danych legacy'],
        ['backend', 'versioning.service.ts cloneVersionData', 'Trzy nowe pola do klonowania wersji',
         'Wytyczna z pamięci projektu: nowe pole z versionId zawsze do klonu', 'Pominięcie = utrata statusów przy nowej wersji wyceny'],
        ['frontend', 'wbsConstants.js', 'JEDEN moduł słownika na trzy osie; usunięcie 4 kopii',
         'Dziś pięć kopii tej samej listy w trzech plikach', 'Dotyka WBSHybridTable, RealizationTab, WbsMaterialsPanel, eksportów'],
        ['frontend', 'kolumna Status w WBS', 'Rozbicie na kolumny: Plan / Zakup / Wykonanie (dwie ostatnie tylko w widoku realizacji)',
         'Etap planu nie ma prawa pokazywać "Na magazynie"', 'Szerokość tabeli — do rozstrzygnięcia w UI'],
        ['frontend', 'RealizationTab', 'Dwa dropdowny zamiast jednego, wg typu liścia', 'Sedno rozdziału zakup / wykonanie',
         'Pozycje material/equipment dostają obie kolumny'],
        ['eksporty', 'PDF / Excel oferty i budżetu', 'Kolumna statusu bierze planStatus; realizacja osobno',
         'Dziś eksport dostaje kod z wymieszanej listy', 'Sprawdzić wbsPdfExport.js i eksport Excela'],
    ],
    [16, 40, 56, 50, 56],
)

# ─────────────────────────────────────────────── 11. Decyzje otwarte
sheet(
    wb, '11. Decyzje otwarte',
    'Do rozstrzygnięcia przed wdrożeniem',
    'Pytania, na które kod nie odpowiada — wymagają Twojej decyzji. Kolumna "Rekomendacja" to propozycja, nie ustalenie.',
    ['#', 'Pytanie', 'Warianty', 'Rekomendacja', 'Konsekwencja wyboru', 'Twoja decyzja'],
    [
        [1, 'Czy oś planu potrzebuje stanów WITHDRAWN i OPTIONAL?',
         'A) tylko 4 podstawowe  B) + Wycofane  C) + Wycofane i Opcja',
         'B — "Wycofane" odróżnia naszą rezygnację od odmowy klienta',
         'Każdy dodatkowy stan to kolejna gałąź w agregacji i w eksporcie', ''],
        [2, 'Gałąź mieszana: PARTIAL czy MIXED?',
         'A) PARTIAL "Częściowo zaakceptowane" + licznik  B) MIXED jak dziś',
         'A — "Mieszany" nie odróżnia 19/20 od 1/20',
         'PARTIAL wymaga liczenia udziału; MIXED jest darmowy', ''],
        [3, 'Czy statusy realizacji mają się wyliczać z LeafActual automatycznie?',
         'A) pełny automat  B) automat tylko jako podpowiedź  C) wyłącznie ręcznie',
         'B — automat podpowiada, człowiek zatwierdza',
         'Pełny automat gubi przypadki, gdy wpis dotyczy zaliczki, nie dostawy', ''],
        [4, 'Czy "Zaakceptowane" ustawia się per pozycja, czy hurtem na akceptacji wersji oferty?',
         'A) per pozycja  B) hurtem na wersji  C) oba',
         'C — hurt na wersji + korekta ręczna na pozycjach spoza zakresu',
         'Hurt wymaga wpięcia w wersjonowanie (ProjectVersion)', ''],
        [5, 'Nocleg i paliwo — oś zakupu czy własna, trzecia ścieżka kosztowa?',
         'A) oś zakupu  B) osobna oś "koszt"',
         'A — kupuje się je tak samo jak materiał, tylko bez magazynu',
         'Wariant B mnoży osie bez zysku informacyjnego', ''],
        [6, 'Usługa obca: jedna oś czy dwie?',
         'A) obie osie (zlecenie + odbiór)  B) tylko wykonanie',
         'A — podwykonawcę się zamawia i osobno odbiera',
         'Wariant B gubi moment podpisania umowy i faktury', ''],
        [7, 'Co ze starym polem WbsNode.status po migracji?',
         'A) usunąć od razu  B) zostawić na jeden cykl wersji  C) zostawić na stałe',
         'B — rollback bez odtwarzania z backupu',
         'Zostawione pole musi mieć zakaz odczytu, inaczej wróci jako trzecie źródło prawdy', ''],
        [8, 'Czy potrzebny ślad zmiany statusu (kto / kiedy / powód)?',
         'A) dwie kolumny na węźle  B) AuditLog  C) brak',
         'B — AuditLog już istnieje i nie puchnie schema',
         'Bez śladu nie da się odtworzyć, kiedy klient odrzucił pozycję', ''],
        [9, 'Czy status wykonania materiału (montaż) prowadzi ekipa, czy kierownik?',
         'A) ekipa z terenu  B) kierownik robót  C) rola per zamówienie',
         'B — spójne z dzisiejszym podziałem uprawnień',
         'Wariant A wymaga uprawnień zapisu dla roli pracownika', ''],
        [10, 'Czy pozycja odrzucona ma znikać z widoku realizacji?',
         'A) znika  B) widoczna wyszarzona  C) filtr',
         'C — filtr domyślnie ukrywa, ale da się pokazać',
         'Znikanie utrudnia wyjaśnienie klientowi, czego nie robimy', ''],
    ],
    [5, 50, 46, 46, 50, 26],
)

wb.save(OUT)
print('OK ->', OUT)
